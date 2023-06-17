import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl.reader.excel import load_workbook

def doZoom(xls_file_path, output_file_path):
    data_frame = pd.read_excel(xls_file_path)
    duration = data_frame['Тривалість']
    Name = data_frame["Ім'я(справжнє)"]
    output_wb = load_workbook(output_file_path)
    output_ws = output_wb.active
    column_duration = "Відвідуваність"
    column_name = "Ім'я(справжнє)"
    target_column_attendance = None
    target_column_name = None
    for column in output_ws.iter_cols(min_row=1, max_row=1):
        if column[0].value == column_duration:
            target_column_attendance = column[0].column
        if column[0].value == column_name:
            target_column_name = column[0].column
        if target_column_attendance != None and target_column_name != None:
            break

    data = zip(duration, Name)

    for i, (attendance, name) in enumerate(data, start=2):
        status = "absent"
        if attendance >= 46:
            status = "present"
        output_ws.cell(row=i, column=target_column_attendance, value=status)
        output_ws.cell(row=i, column=target_column_name, value=name)

    output_wb.save(output_file_path)
    print("Success")

def doMoodle(xls_file_path, output_file_path):
    data_frame = pd.read_excel(xls_file_path)
    Grades = data_frame['Загальне за курс (Бали)']
    Surname = data_frame['Прізвище']
    Name = data_frame["Ім'я"]
    output_wb = load_workbook(output_file_path)
    output_ws = output_wb.active
    column_grade = "Бали"
    column_surname = "Прізвище"
    column_name = "Ім'я"
    target_column_grade = None
    target_column_name = None
    target_column_surname = None
    for column in output_ws.iter_cols(min_row=1, max_row=1):
        if column[0].value == column_grade:
            target_column_grade = column[0].column
        if column[0].value == column_surname:
            target_column_surname = column[0].column
        if column[0].value == column_name:
            target_column_name = column[0].column
        if target_column_surname!=None and target_column_name!=None and target_column_grade!=None:
            break

    data = zip(Grades, Surname, Name)
    for i, (grade, surname, name) in enumerate(data, start=2):
        output_ws.cell(row=i, column=target_column_grade, value=grade)
        output_ws.cell(row=i, column=target_column_surname, value=surname)
        output_ws.cell(row=i, column=target_column_name, value=name)

    output_wb.save(output_file_path)
    print("Success")

def select_input_file():
    global input_file_path
    input_file_path = filedialog.askopenfilename(title="Select input file",
                                                 filetypes=(("Excel files", "*.xls *.xlsx"),("All files", "*.*")))
    if input_file_path:
        if not (input_file_path.endswith(".xls") or input_file_path.endswith(".xlsx")):
            messagebox.showerror("Error", "Input file must be a .xls or .xlsx file!")
            input_file_path = ""
        else:
            input_file_label['text'] = input_file_path

def select_output_file():
    global output_file_path
    output_file_path = filedialog.asksaveasfilename(title="Select output file", defaultextension=".xls",
                                                    filetypes=(("Excel files", "*.xls *.xlsx"),("All files", "*.*")))
    if output_file_path:
        if not (output_file_path.endswith(".xls") or output_file_path.endswith(".xlsx")):
            messagebox.showerror("Error", "Output file must be a .xls or .xlsx file!")
            output_file_path = ""
        else:
            output_file_label['text'] = output_file_path

def process_files():
    if input_file_path and output_file_path:
        if "Zoom" in input_file_path:
            doZoom(input_file_path, output_file_path)
        if "Moodle" in input_file_path:
            doMoodle(input_file_path, output_file_path)

root = tk.Tk()
root.title("KSE Automatization")

title_label = tk.Label(root, text="KSE Automatization", font=('Verdana', 18, 'bold'))
title_label.pack(pady=10)

explanation_label = tk.Label(root, text="This program helps you systematize tables.", font=('Verdana', 12))
explanation_label.pack(pady=10)

input_file_button = tk.Button(root, text="Select Input File", command=select_input_file, font=('Verdana', 14), padx=20, pady=10)
input_file_button.pack()

input_file_label = tk.Label(root, text="", font=('Verdana', 10))
input_file_label.pack(pady=10)

output_file_button = tk.Button(root, text="Select Output File", command=select_output_file, font=('Verdana', 14), padx=20, pady=10)
output_file_button.pack()

output_file_label = tk.Label(root, text="", font=('Verdana', 10))
output_file_label.pack(pady=10)

process_button = tk.Button(root, text="Process Files", command=process_files, font=('Verdana', 14), padx=20, pady=10)
process_button.pack()

root.mainloop()